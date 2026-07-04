#!/usr/bin/env python3
"""
build_paper_tables.py
  Table#1 / Table#4 를 이미지 레이아웃 그대로(2단 헤더 + 좌측 그룹 병합 + main 셀 박스)
  HTML 로 생성한다. 값은 result/ 아래 summary_metrics.csv 에서 채우고, 없으면 빈 칸.

  출력: tables/table1.html, tables/table4.html
"""
from __future__ import annotations

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

# summary_metrics.csv Metric 이름 -> 표시 열
METRIC_MAP = [
    ("Loc.Acc", "Detection Acc"),
    ("Cls.Acc", "Classification Acc"),
    ("Overall.Ac", "Overall Acc"),
    ("mAP@0.5", "mAP@0.5"),
    ("precision", "Precision"),
    ("recall", "Recall"),
]
METRIC_COLS = [m[0] for m in METRIC_MAP]


def read_summary(eval_dir: Path) -> dict[str, str] | None:
    csv_path = eval_dir / "summary_metrics.csv"
    if not csv_path.exists():
        return None
    data: dict[str, str] = {}
    with csv_path.open(encoding="utf-8") as fh:
        for row in csv.reader(fh):
            if len(row) >= 2:
                data[row[0]] = row[1]
    return data


def fmt(v: str | None) -> str:
    if v is None or v == "" or v == "N/A":
        return ""
    try:
        return f"{float(v):.3f}"
    except ValueError:
        return str(v)


def resolve_eval_dir(mode_root: str, classtag: str, train: str, model: str, test: str) -> Path | None:
    """(train,model,test) 조합의 eval 디렉터리를 찾아 반환. 없으면 None."""
    if not model:  # zero-shot 등 결과 없음
        return None
    if train == test:  # self-test
        base = ROOT / "result" / mode_root / test
        cands = [
            base / f"{model}_{classtag}_eval",
            base / f"{model}_{classtag}_prediction" / "evaluation",  # dino 계열
        ]
    elif train == "merge":
        base = ROOT / "result" / mode_root / "merge" / test
        cands = [
            base / f"{model}_{classtag}_{test}_eval",
            base / f"{model}_{classtag}_{test}_prediction" / "evaluation",
        ]
    else:
        return None
    for c in cands:
        if (c / "summary_metrics.csv").exists():
            return c
    return None


def cell_values(mode_root: str, classtag: str, train: str, model: str, test: str) -> list[str]:
    ed = resolve_eval_dir(mode_root, classtag, train, model, test)
    if ed is None:
        return [""] * len(METRIC_MAP)
    data = read_summary(ed)
    if data is None:
        return [""] * len(METRIC_MAP)
    return [fmt(data.get(src)) for _, src in METRIC_MAP]


def build_rows(blocks, trains, testcols, models_by_block) -> list[list[str]]:
    """이미지 레이아웃(2단 헤더 + block/train/method + test그룹×6지표)을 CSV 행으로."""
    rows: list[list[str]] = []
    # 헤더 1: test 데이터 그룹명 (6칸마다 반복)
    h1 = ["", "", ""]
    for name, _k, _cap in testcols:
        h1 += [name] + [""] * (len(METRIC_COLS) - 1)
    rows.append(h1)
    # 헤더 2: 그룹 설명
    h2 = ["", "", ""]
    for _n, _k, cap in testcols:
        h2 += [cap] + [""] * (len(METRIC_COLS) - 1)
    rows.append(h2)
    # 헤더 3: block/train/method + 지표명
    h3 = ["block", "train", "method"]
    for _ in testcols:
        h3 += METRIC_COLS
    rows.append(h3)

    # 데이터
    for block_label, mode_root, classtag in blocks:
        models = models_by_block[block_label]
        for train_name, train_key in trains:
            for disp, model_key in models:
                row = [block_label, train_name, disp]
                for _tn, test_key, _cap in testcols:
                    row += cell_values(mode_root, classtag, train_key, model_key, test_key)
                rows.append(row)
    return rows


# ---------------- Table#1 ----------------
T1_TESTCOLS = [
    ("TomatOD", "tomatod", "Varying illumination or darkness"),
    ("Laboro Big", "big", "Standard tomato instances"),
    ("Laboro Little", "little", "Cherry tomato instances"),
]
T1_TRAINS = [
    ("TomatOD", "tomatod"),
    ("Laboro Big", "big"),
    ("Laboro Little", "little"),
    ("Merge", "merge"),
]
T1_MODELS = [
    ("YOLO11l", "yolo11"),
    ("YOLO12l", "yolo12"),
    ("YOLO World(FT)", "yolo_world"),
    ("RT-DETR", "rt_detr"),
    ("RF-DETR", "rf_detr"),
    ("Grounding DINO", "grounding_dino"),
    ("DINO-DETR", "dino"),
]
T1_BLOCKS = [
    ("3cls", "detection_only", "3cls"),
    ("1cls + VLM (DaR)", "detection_reasoning", "1cls"),
]

# ---------------- Table#4 ----------------
T4_TESTCOLS = [
    ("Rho2Pheno", "rho2pheno", "Unseen: smartfarm occlussioned dataset"),
    ("Custom Tomato", "custom_tomato", "Unseen: smartfarm occlussioned dataset"),
]
T4_TRAINS = [
    ("Rho2Pheno", "rho2pheno"),
    ("Custom Tomato", "custom_tomato"),
]
T4_MODELS_2CLS = [
    ("YOLO11l", "yolo11"),
    ("YOLO12l", "yolo12"),
    ("YOLO World(FT)", "yolo_world"),
    ("RT-DETR", "rt_detr"),
    ("RF-DETR", "rf_detr"),
    ("YOLO World(ZS)", "yolo_world_zs"),
    ("Grounding DINO", "grounding_dino"),
    ("Owl-ViT", "owl_vit"),
]
T4_MODELS_1CLS = [
    ("YOLO11l", "yolo11"),
    ("YOLO12l", "yolo12"),
    ("YOLO World(FT)", "yolo_world"),
    ("RT-DETR", "rt_detr"),
    ("RF-DETR", "rf_detr"),
]
T4_BLOCKS = [
    ("2cls", "detection_only", "2cls"),
    ("1cls + VLM (DaR)", "detection_reasoning", "1cls"),
]

def write_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        csv.writer(fh).writerows(rows)


def _num(v: str):
    try:
        return float(v)
    except ValueError:
        return v or None


def write_xlsx(path: Path, blocks, trains, testcols, models_by_block) -> None:
    """이미지처럼 병합 헤더 + 색상/박스 강조가 들어간 xlsx 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    nm = len(METRIC_COLS)
    fill_grp = PatternFill("solid", fgColor="FDE9D9")
    fill_cap = PatternFill("solid", fgColor="FFF2CC")
    fill_metric = PatternFill("solid", fgColor="F2F2F2")
    fill_block = PatternFill("solid", fgColor="DBE5F1")
    fill_train = PatternFill("solid", fgColor="FDEADA")
    fill_box = PatternFill("solid", fgColor="DCE6F7")  # main 셀 강조
    thin = Side(style="thin", color="BBBBBB")
    box_side = Side(style="medium", color="1F3864")
    border_thin = Border(thin, thin, thin, thin)
    border_box = Border(box_side, box_side, box_side, box_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    zs_font = Font(color="C00000", bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "table"

    ncols = 3 + nm * len(testcols)
    # row1 group, row2 caption
    for i, (name, _k, cap) in enumerate(testcols):
        c0 = 4 + i * nm
        ws.cell(1, c0, name)
        ws.cell(2, c0, cap)
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + nm - 1)
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + nm - 1)
    for r, fill in ((1, fill_grp), (2, fill_cap)):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.alignment = center
            cell.border = border_thin
    # row3
    hdr3 = ["block", "train", "method"] + METRIC_COLS * len(testcols)
    for c, val in enumerate(hdr3, start=1):
        cell = ws.cell(3, c, val)
        cell.fill = fill_metric
        cell.font = bold
        cell.alignment = center
        cell.border = border_thin

    # 데이터
    r = 4
    for block_label, mode_root, classtag in blocks:
        models = models_by_block[block_label]
        block_start = r
        for train_name, train_key in trains:
            train_start = r
            for disp, model_key in models:
                bcell = ws.cell(r, 1, block_label)
                tcell = ws.cell(r, 2, train_name)
                mcell = ws.cell(r, 3, disp)
                bcell.fill, tcell.fill = fill_block, fill_train
                bcell.font = tcell.font = bold
                for cc in (bcell, tcell, mcell):
                    cc.alignment = center
                    cc.border = border_thin
                zs = model_key == "" or model_key.endswith("_zs") or model_key in {"grounding_dino", "owl_vit"}
                if zs:
                    mcell.font = zs_font
                for i, (_tn, test_key, _cap) in enumerate(testcols):
                    main = (train_key == test_key) or (train_key == "merge")
                    vals = cell_values(mode_root, classtag, train_key, model_key, test_key)
                    for j, v in enumerate(vals):
                        cell = ws.cell(r, 4 + i * nm + j, _num(v))
                        cell.alignment = center
                        if main:
                            cell.fill = fill_box
                            cell.border = border_box
                        else:
                            cell.border = border_thin
                r += 1
            if r - train_start > 1:
                ws.merge_cells(start_row=train_start, start_column=2, end_row=r - 1, end_column=2)
        if r - block_start > 1:
            ws.merge_cells(start_row=block_start, start_column=1, end_row=r - 1, end_column=1)

    # 열 너비
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 15
    for c in range(4, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8
    ws.freeze_panes = "D4"
    wb.save(path)


def main() -> None:
    t1 = build_rows(
        T1_BLOCKS, T1_TRAINS, T1_TESTCOLS,
        {"3cls": T1_MODELS, "1cls + VLM (DaR)": T1_MODELS},
    )
    write_csv(ROOT / "table1.csv", t1)

    t4 = build_rows(
        T4_BLOCKS, T4_TRAINS, T4_TESTCOLS,
        {"2cls": T4_MODELS_2CLS, "1cls + VLM (DaR)": T4_MODELS_1CLS},
    )
    write_csv(ROOT / "table4.csv", t4)

    t1_models = {"3cls": T1_MODELS, "1cls + VLM (DaR)": T1_MODELS}
    write_xlsx(ROOT / "table1.xlsx", T1_BLOCKS, T1_TRAINS, T1_TESTCOLS, t1_models)
    t4_models = {"2cls": T4_MODELS_2CLS, "1cls + VLM (DaR)": T4_MODELS_1CLS}
    write_xlsx(ROOT / "table4.xlsx", T4_BLOCKS, T4_TRAINS, T4_TESTCOLS, t4_models)

    for f in ("table1.csv", "table4.csv", "table1.xlsx", "table4.xlsx"):
        print("wrote:", ROOT / f)


if __name__ == "__main__":
    main()
